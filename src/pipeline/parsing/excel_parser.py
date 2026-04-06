"""Excel 파서.

Excel (.xlsx, .xls) → Markdown 테이블 변환.
멀티시트를 각각 독립 마크다운 섹션으로 변환한다.
"""

from __future__ import annotations

import time

from src.observability.logging import get_logger
from src.pipeline.parsing.base import FormatParser, ParseMetrics, ParseResult
from src.pipeline.parsing.metrics import compute_markdown_metrics

logger = get_logger(__name__)

_SUPPORTED = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
]

_MAX_ROWS_PER_SHEET = 10000
_MAX_SHEETS = 20


class ExcelParser:
    """Excel → Markdown 테이블 변환. 멀티시트 지원."""

    def __init__(self, max_rows_per_sheet: int = _MAX_ROWS_PER_SHEET):
        self._max_rows = max_rows_per_sheet

    async def parse(self, file_bytes: bytes, file_name: str = "") -> ParseResult:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing: pip install openpyxl")

        t0 = time.time()

        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        sections: list[str] = []
        total_rows = 0
        sheet_count = 0

        for sheet_name in wb.sheetnames[:_MAX_SHEETS]:
            ws = wb[sheet_name]
            rows: list[list[str]] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= self._max_rows:
                    break
                cells = [str(c) if c is not None else "" for c in row]
                # 완전히 빈 행 스킵
                if any(c.strip() for c in cells):
                    rows.append(cells)

            if not rows:
                continue

            sheet_count += 1
            total_rows += len(rows)
            section = _sheet_to_markdown(sheet_name, rows)
            sections.append(section)

        wb.close()

        markdown = "\n\n".join(sections)
        elapsed_ms = (time.time() - t0) * 1000

        metrics = compute_markdown_metrics(
            markdown=markdown,
            parse_time_ms=elapsed_ms,
            total_pages=sheet_count,
            parser_used="excel",
            parser_reason=f"sheets={sheet_count}, total_rows={total_rows}",
        )

        logger.info(
            "excel_parsed",
            file_name=file_name,
            sheets=sheet_count,
            total_rows=total_rows,
            latency_ms=round(elapsed_ms, 1),
        )

        return ParseResult(
            markdown=markdown,
            metrics=metrics,
            source_mime_type=_SUPPORTED[0],
            source_file_name=file_name,
        )

    def supported_mime_types(self) -> list[str]:
        return list(_SUPPORTED)


def _sheet_to_markdown(sheet_name: str, rows: list[list[str]]) -> str:
    """시트 데이터를 마크다운 테이블 섹션으로 변환한다."""
    if not rows:
        return ""

    header = rows[0]
    data_rows = rows[1:]
    col_count = len(header)

    parts: list[str] = [f"## {sheet_name}\n"]

    # 헤더에 빈 셀이 많으면 자동 열 이름 생성
    header_cells = []
    for i, h in enumerate(header):
        cell = h.strip()
        if not cell:
            cell = f"Column_{i + 1}"
        header_cells.append(_escape_pipe(cell))

    parts.append("| " + " | ".join(header_cells) + " |")
    parts.append("| " + " | ".join(["---"] * col_count) + " |")

    for row in data_rows:
        # 열 수 정규화
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        elif len(row) > col_count:
            row = row[:col_count]
        cells = [_escape_pipe(c) for c in row]
        parts.append("| " + " | ".join(cells) + " |")

    return "\n".join(parts)


def _escape_pipe(value: str) -> str:
    """마크다운 테이블 셀 이스케이프."""
    return value.replace("|", "\\|").replace("\n", " ").strip()
